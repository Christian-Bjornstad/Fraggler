import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook

from core.qc.qc_excel import apply_pk_excel_styling


class TestQcExcelStyling(unittest.TestCase):
    def _count_conditional_rules(self, path: Path) -> int:
        wb = load_workbook(path)
        total = 0
        for sheet_name in ("PK_Runs", "PK_Markers"):
            ws = wb[sheet_name]
            total += sum(len(rules) for rules in ws.conditional_formatting._cf_rules.values())
        return total

    def test_apply_pk_excel_styling_replaces_existing_rules(self):
        with TemporaryDirectory() as tmp:
            excel_path = Path(tmp) / "qc.xlsx"

            wb = Workbook()
            ws_runs = wb.active
            ws_runs.title = "PK_Runs"
            ws_runs.append(
                [
                    "run_key",
                    "run_date",
                    "run_code",
                    "pcr_date",
                    "file",
                    "control",
                    "assay",
                    "well",
                    "batch",
                    "ladder",
                    "ladder_qc",
                    "ladder_r2",
                    "bp_min",
                    "bp_max",
                ]
            )
            ws_runs.append(["run-1", "", "", "", "file.fsa", "PK", "FLT3", "", "", "ROX400HD", "OK", 0.998, 50, 1000])

            ws_markers = wb.create_sheet("PK_Markers")
            ws_markers.append(
                [
                    "run_key",
                    "run_date",
                    "run_code",
                    "pcr_date",
                    "file",
                    "control",
                    "assay",
                    "well",
                    "batch",
                    "marker_name",
                    "kind",
                    "channel",
                    "expected_bp",
                    "window_bp",
                    "ok",
                    "found_bp",
                    "delta_bp",
                    "height",
                    "area",
                    "reason",
                ]
            )
            ws_markers.append(
                [
                    "run-1",
                    "",
                    "",
                    "",
                    "file.fsa",
                    "PK",
                    "FLT3",
                    "",
                    "",
                    "marker-1",
                    "sample",
                    "DATA1",
                    100.0,
                    1.0,
                    True,
                    100.2,
                    0.2,
                    123.0,
                    456.0,
                    "",
                ]
            )
            wb.save(excel_path)

            apply_pk_excel_styling(excel_path)
            first_count = self._count_conditional_rules(excel_path)

            apply_pk_excel_styling(excel_path)
            second_count = self._count_conditional_rules(excel_path)

        self.assertEqual(first_count, 7)
        self.assertEqual(second_count, 7)


if __name__ == "__main__":
    unittest.main()
