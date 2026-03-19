"""
Fraggler QC — Excel trend tracking (append-mode).
"""
from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from core.qc.qc_rules import QCRules
from core.qc.qc_markers import (
    markers_for_entry,
    find_peak_near_bp,
    parse_run_code_from_filename,
    parse_pcr_date_from_filename,
    parse_well_from_filename,
    parse_batch_from_filename,
    control_id_from_filename,
    ladder_qc_grade,
    make_run_key,
)


def update_excel_trends(excel_path: Path, entries: list[dict], rules: QCRules, run_ts: str):
    """
    Oppdaterer Excel automatisk (append).
    Kun PK/PK1/PK2 spores (fokus på positive kontroller).
    Lager/oppdaterer to ark:
      - PK_Runs
      - PK_Markers
    """
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    rows_runs = []
    rows_markers = []

    for e in entries:
        fsa = e["fsa"]
        fname = fsa.file_name

        ctrl = control_id_from_filename(fname)
        if ctrl not in {"PK", "PK1", "PK2"}:
            continue  # <-- kun PK

        assay = e.get("assay")
        ladder = e.get("ladder")
        r2 = e.get("ladder_r2")
        grade, note = ladder_qc_grade(r2, rules)

        pcr_date = parse_pcr_date_from_filename(fname)     # yyyy-mm-dd eller None
        well = parse_well_from_filename(fname)             # f.eks. G09 eller None
        batch = parse_batch_from_filename(fname)           # f.eks. C991475U eller None


        run_date = parse_pcr_date_from_filename(fname)   # YYYY-MM-DD
        run_code = parse_run_code_from_filename(fname)   # H9C0U3SI etc.
        run_key = make_run_key(fname)                    # YYYY-MM-DD_H9C0U3SI


        # --------- PK_Runs (én rad per fil) ----------
        rows_runs.append({
            "run_key": run_key,
            "run_date": run_date,
            "run_code": run_code,
            "pcr_date": pcr_date,
            "file": fname,
            "control": ctrl,
            "assay": assay,
            "well": well,
            "batch": batch,
            "ladder": e.get("ladder"),
            "qc_grade": grade,
            "ladder_r2": None if r2 is None or not np.isfinite(r2) else float(r2),
            "bp_min": float(e.get("bp_min", np.nan)),
            "bp_max": float(e.get("bp_max", np.nan)),
        })

        # --------- PK_Markers (én rad per markør) ----------
        # (markører fylles kun for PK i din marker-config, så dette blir pent)
        mrs = e.get("qc_marker_results", []) or []
        for mr in mrs:
            rows_markers.append({
                "run_key": run_key,
                "run_date": run_date,
                "run_code": run_code,
                "pcr_date": pcr_date,
                "file": fname,
                "control": ctrl,
                "assay": assay,
                "well": well,
                "batch": batch,

                "marker_name": mr.get("name"),
                "kind": mr.get("kind"),           # sample/ladder
                "channel": mr.get("channel"),     # DATA1/DATA2/DATA3/DATA105/DATA4
                "expected_bp": float(mr.get("expected_bp", np.nan)),
                "window_bp": float(mr.get("window_bp", np.nan)),
                "ok": bool(mr.get("ok", False)),

                "found_bp": None if not mr.get("ok") else float(mr.get("found_bp")),
                "delta_bp": None if not mr.get("ok") else float(mr.get("found_bp")) - float(mr.get("expected_bp")),
                "height": None if not mr.get("ok") else float(mr.get("height")),
                "area": None if not mr.get("ok") else float(mr.get("area")),
                "reason": None if mr.get("ok") else mr.get("reason"),
            })

    df_runs = pd.DataFrame(rows_runs)
    df_markers = pd.DataFrame(rows_markers)

    # Fjern interne duplikater i denne kjøringen
    if not df_runs.empty:
        df_runs = df_runs.drop_duplicates(subset=["file"], keep="last")
    if not df_markers.empty:
        df_markers = df_markers.drop_duplicates(subset=["file", "marker_name"], keep="last")



    # Lokal dedupe (inni samme kjøring)
    if not df_runs.empty:
        df_runs = df_runs.drop_duplicates(subset=["file"], keep="last")
    if not df_markers.empty:
        df_markers = df_markers.drop_duplicates(subset=["file", "marker_name"], keep="last")

    # Global dedupe (Excel-historikk + nye rader)
    if excel_path.exists():
        try:
            xls = pd.ExcelFile(excel_path, engine="openpyxl")
            has_runs = "PK_Runs" in xls.sheet_names
            has_markers = "PK_Markers" in xls.sheet_names
        except Exception:
            has_runs = False
            has_markers = False

        old_runs = pd.read_excel(excel_path, sheet_name="PK_Runs", engine="openpyxl") if has_runs else pd.DataFrame()
        old_markers = pd.read_excel(excel_path, sheet_name="PK_Markers", engine="openpyxl") if has_markers else pd.DataFrame()

        all_runs = pd.concat([old_runs, df_runs], ignore_index=True)
        all_markers = pd.concat([old_markers, df_markers], ignore_index=True)

        # Global dedupe-nøkler:
        # - Runs: file (én rad per fil)
        # - Markers: file + marker_name (én rad per markør per fil)
        if not all_runs.empty and "file" in all_runs.columns:
            all_runs = all_runs.drop_duplicates(subset=["file"], keep="last")
        if not all_markers.empty and {"file", "marker_name"}.issubset(all_markers.columns):
            all_markers = all_markers.drop_duplicates(subset=["file", "marker_name"], keep="last")

        # Skriv tilbake (replace) for å unngå at Excel vokser ukontrollert
        with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            all_runs.to_excel(writer, sheet_name="PK_Runs", index=False)
            all_markers.to_excel(writer, sheet_name="PK_Markers", index=False)

    else:
        # Ny fil
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df_runs.to_excel(writer, sheet_name="PK_Runs", index=False)
            df_markers.to_excel(writer, sheet_name="PK_Markers", index=False)

    # Styling (farger)
    apply_pk_excel_styling(excel_path)

def apply_pk_excel_styling(excel_path: Path):
    from openpyxl import load_workbook
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(excel_path)

    fill_ok = PatternFill("solid", fgColor="E6F4EA")   # grønn
    fill_warn = PatternFill("solid", fgColor="FFF4E5") # gul
    fill_fail = PatternFill("solid", fgColor="FDE7E9") # rød

    def _reset_conditional_formatting(ws) -> None:
        """Replace existing conditional-format rules instead of piling on duplicates."""
        cf_rules = getattr(ws.conditional_formatting, "_cf_rules", None)
        if cf_rules is not None:
            cf_rules.clear()

    # ---- PK_Runs: ladder_qc farger ----
    if "PK_Runs" in wb.sheetnames:
        ws = wb["PK_Runs"]
        _reset_conditional_formatting(ws)
        headers = {c.value: c.column for c in ws[1] if c.value}
        if "ladder_qc" in headers:
            col = headers["ladder_qc"]
            L = get_column_letter(col)
            rng = f"{L}2:{L}{ws.max_row}"
            ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"OK"'], fill=fill_ok))
            ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"WARN"'], fill=fill_warn))
            ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"FAIL"'], fill=fill_fail))

    # ---- PK_Markers: delta_bp farger (start: +/-0.5 gul, +/-1.0 rød) ----
    if "PK_Markers" in wb.sheetnames:
        ws = wb["PK_Markers"]
        _reset_conditional_formatting(ws)
        headers = {c.value: c.column for c in ws[1] if c.value}
        if "delta_bp" in headers:
            col = headers["delta_bp"]
            L = get_column_letter(col)
            rng = f"{L}2:{L}{ws.max_row}"

            # rød hvis delta >= 1.0 eller <= -1.0
            ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["1.0"], fill=fill_fail))
            ws.conditional_formatting.add(rng, CellIsRule(operator="lessThanOrEqual", formula=["-1.0"], fill=fill_fail))

            # gul hvis delta >= 0.5 eller <= -0.5
            ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["0.5"], fill=fill_warn))
            ws.conditional_formatting.add(rng, CellIsRule(operator="lessThanOrEqual", formula=["-0.5"], fill=fill_warn))

    wb.save(excel_path)
