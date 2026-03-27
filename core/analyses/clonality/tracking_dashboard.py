from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="DCEAF7")
SUBHEADER_FILL = PatternFill("solid", fgColor="EAF2F8")
CARD_BLUE = PatternFill("solid", fgColor="D9EAF7")
CARD_TEAL = PatternFill("solid", fgColor="DDF4F1")
CARD_ORANGE = PatternFill("solid", fgColor="FCE4D6")
CARD_RED = PatternFill("solid", fgColor="FBE5E7")
CARD_GOLD = PatternFill("solid", fgColor="FFF2CC")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN_GRAY = Side(style="thin", color="D9E2F2")
BOX_BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
DASHBOARD_SHEETS = [
    "Dashboard",
    "Dashboard_Data",
    "Assay_Summary",
    "Run_Summary",
    "Control_Summary",
    "Daily_Trend",
    "PK_Sample_Delta",
    "PK_Ladder_Delta",
    "PK_Outliers",
    "Review_Files",
]


def _require_cols(cols: dict[str, str], *names: str) -> dict[str, str]:
    missing = [name for name in names if name not in cols]
    if missing:
        raise KeyError(f"Missing required columns: {', '.join(missing)}")
    return {name: cols[name] for name in names}


def _range_ref(sheet_name: str, col: str, *, start_row: int = 2, end_row: int | None = None) -> str:
    end = end_row if end_row is not None else 1048576
    return f"{sheet_name}!${col}${start_row}:${col}${end}"


def _count_nonblank_formula(sheet_name: str, col: str) -> str:
    return f'=COUNTIF({_range_ref(sheet_name, col)}, "<>")'


def _count_review_formula(sheet_name: str, cols: dict[str, str], row_ref: str) -> str:
    return (
        f'=COUNTIFS({_range_ref(sheet_name, cols["LadderQC"])}, "<>", '
        f'{_range_ref(sheet_name, cols["LadderQC"])}, "<>ok", '
        f'{_range_ref(sheet_name, cols["Assay"])}, {row_ref})'
    )


def refresh_clonality_tracking_dashboard(excel_path: Path) -> None:
    if not excel_path.exists():
        return

    with pd.ExcelFile(excel_path, engine="openpyxl") as xls:
        required = {"Patient_Runs", "Control_Runs", "PK_Peaks"}
        if not required.issubset(set(xls.sheet_names)):
            return

    patient = pd.read_excel(excel_path, sheet_name="Patient_Runs", engine="openpyxl").fillna("")
    control = pd.read_excel(excel_path, sheet_name="Control_Runs", engine="openpyxl").fillna("")
    pk = pd.read_excel(excel_path, sheet_name="PK_Peaks", engine="openpyxl").fillna("")

    patient["RunDate"] = pd.to_datetime(patient.get("RunDate"), errors="coerce")
    control["RunDate"] = pd.to_datetime(control.get("RunDate"), errors="coerce")
    pk["RunDate"] = pd.to_datetime(pk.get("RunDate"), errors="coerce")
    pk["AbsDeltaBP"] = pd.to_numeric(pk.get("AbsDeltaBP"), errors="coerce")
    pk["DeltaBP"] = pd.to_numeric(pk.get("DeltaBP"), errors="coerce")

    assay_series = [s for s in [patient.get("Assay"), control.get("Assay"), pk.get("Assay")] if s is not None]
    assays = sorted({str(v).strip() for v in pd.concat(assay_series, ignore_index=True).tolist() if str(v).strip()}) if assay_series else []
    controls = sorted({str(v).strip() for v in control.get("Control", pd.Series(dtype=str)).tolist() if str(v).strip()})
    review_pairs = (
        control[["Control", "Assay"]]
        .drop_duplicates()
        .sort_values(["Control", "Assay"])
        .values.tolist()
    )
    run_days = sorted(
        {
            d.strftime("%Y-%m-%d")
            for d in pd.concat([patient["RunDate"], control["RunDate"]]).dropna().tolist()
        }
    )

    pk_outlier_columns = [c for c in ["IdentityKey", "Assay", "Control", "MarkerName", "ExpectedBP", "FoundBP", "DeltaBP", "AbsDeltaBP", "Height", "Reason"] if c in pk.columns]
    pk_sample_outliers = (
        pk.loc[
            (pk.get("Kind", "").astype(str).str.lower() == "sample")
            & (pd.to_numeric(pk.get("AbsDeltaBP"), errors="coerce") > 2),
            pk_outlier_columns,
        ]
        .copy()
        .sort_values([c for c in ["AbsDeltaBP", "Assay", "IdentityKey"] if c in pk_outlier_columns], ascending=[False, True, True][: len([c for c in ["AbsDeltaBP", "Assay", "IdentityKey"] if c in pk_outlier_columns])])
        .head(50)
    )
    run_frames = [frame for frame in [patient.assign(Scope="Patient"), control.assign(Scope="Control")] if not frame.empty]
    review_frame = pd.concat(run_frames, ignore_index=True) if run_frames else pd.DataFrame()
    review_columns = [c for c in ["Scope", "IdentityKey", "Control", "Assay", "SourceRunDir", "RunDate", "LadderQC", "LadderFitStrategy", "LadderExpectedStepCount", "LadderFittedStepCount", "LadderR2"] if c in review_frame.columns]
    review_files = (
        review_frame
        .loc[
            lambda df: df["LadderQC"].astype(str).str.strip().str.lower().ne("ok") & df["LadderQC"].astype(str).str.strip().ne(""),
            review_columns,
        ]
        .copy()
        .sort_values([c for c in ["Scope", "Assay", "RunDate", "IdentityKey"] if c in review_columns])
    ) if run_frames else pd.DataFrame(columns=review_columns)
    if not review_files.empty:
        review_files["RunDate"] = pd.to_datetime(review_files["RunDate"], errors="coerce").dt.strftime("%Y-%m-%d")

    wb = load_workbook(excel_path)
    _ensure_abs_delta_column(wb["PK_Peaks"])
    for name in DASHBOARD_SHEETS:
        if name in wb.sheetnames:
            del wb[name]

    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    dashboard = wb.create_sheet("Dashboard", 0)
    data_ws = wb.create_sheet("Dashboard_Data")
    data_ws.sheet_state = "hidden"
    assay_ws = wb.create_sheet("Assay_Summary")
    run_ws = wb.create_sheet("Run_Summary")
    control_ws = wb.create_sheet("Control_Summary")
    daily_ws = wb.create_sheet("Daily_Trend")
    pk_sample_ws = wb.create_sheet("PK_Sample_Delta")
    pk_ladder_ws = wb.create_sheet("PK_Ladder_Delta")
    pk_outlier_ws = wb.create_sheet("PK_Outliers")
    review_ws = wb.create_sheet("Review_Files")

    patient_cols = _col_map(wb["Patient_Runs"])
    control_cols = _col_map(wb["Control_Runs"])
    pk_cols = _col_map(wb["PK_Peaks"])
    patient_last = max(wb["Patient_Runs"].max_row, 2)
    control_last = max(wb["Control_Runs"].max_row, 2)
    pk_last = max(wb["PK_Peaks"].max_row, 2)

    _write_helper_lists(data_ws, assays, controls, review_pairs, run_days)
    _build_assay_summary(assay_ws, assays, patient_cols, control_cols, patient_last, control_last)
    _build_run_summary(run_ws, assays, patient_cols, patient_last)
    _build_control_summary(control_ws, review_pairs, control_cols, control_last)
    _build_daily_trend(daily_ws, run_days, patient_cols, control_cols)
    _build_pk_summary(
        pk_sample_ws,
        assays,
        pk_cols,
        pk_last,
        kind="sample",
        title="PK Sample Delta",
    )
    _build_pk_summary(
        pk_ladder_ws,
        assays,
        pk_cols,
        pk_last,
        kind="ladder",
        title="PK Ladder Delta",
    )
    _write_static_frame(pk_outlier_ws, "PK Outliers", pk_sample_outliers)
    _write_static_frame(review_ws, "Review Files", review_files)
    _build_dashboard(
        dashboard,
        len(assays),
        len(run_days),
        len(review_files),
        len(pk_sample_outliers),
        list(pk_sample_outliers.columns),
        assay_ws,
        control_ws,
        pk_sample_ws,
        pk_ladder_ws,
        daily_ws,
    )

    wb.save(excel_path)


def _ensure_abs_delta_column(ws) -> None:
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    if "AbsDeltaBP" in headers:
        return

    delta_idx = headers.index("DeltaBP") + 1 if "DeltaBP" in headers else None
    next_col = ws.max_column + 1
    ws.cell(1, next_col, "AbsDeltaBP")
    for row_idx in range(2, ws.max_row + 1):
        delta_value = ws.cell(row_idx, delta_idx).value if delta_idx else None
        if delta_value in ("", None):
            continue
        try:
            ws.cell(row_idx, next_col, abs(float(delta_value)))
        except (TypeError, ValueError):
            continue


def _col_map(ws) -> dict[str, str]:
    return {str(cell.value): get_column_letter(cell.column) for cell in ws[1] if cell.value}


def _style_table(ws, header_row: int, data_end_row: int, start_col: int, end_col: int) -> None:
    for cell in ws[header_row]:
        if start_col <= cell.column <= end_col:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BOX_BORDER
    for row in ws.iter_rows(min_row=header_row + 1, max_row=max(data_end_row, header_row + 1), min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = BOX_BORDER
            if cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FBFF")


def _autofit(ws) -> None:
    for column_cells in ws.columns:
        length = 0
        col_idx = column_cells[0].column
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, len(value))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(length + 2, 10), 30)


def _write_helper_lists(ws, assays: list[str], controls: list[str], review_pairs: list[list[str]], run_days: list[str]) -> None:
    ws["A1"] = "Assay"
    ws["E1"] = "Control"
    ws["F1"] = "Assay"
    ws["J1"] = "RunDay"
    for row, assay in enumerate(assays, start=2):
        ws.cell(row, 1, assay)
    for row, control in enumerate(controls, start=2):
        ws.cell(row, 5, control)
    for row, pair in enumerate(review_pairs, start=2):
        ws.cell(row, 5, pair[0])
        ws.cell(row, 6, pair[1])
    for row, run_day in enumerate(run_days, start=2):
        ws.cell(row, 10, run_day)


def _build_assay_summary(ws, assays: list[str], patient_cols: dict[str, str], control_cols: dict[str, str], patient_last: int, control_last: int) -> None:
    ws.sheet_properties.tabColor = "2F75B5"
    headers = ["Assay", "Files", "PatientFiles", "ControlFiles", "LadderReview", "AvgR2", "PartialFits", "ReviewRate"]
    ws.append(headers)
    assay_col = "$A2"
    patient_req = _require_cols(patient_cols, "IdentityKey", "Assay", "LadderQC", "LadderR2", "LadderExpectedStepCount", "LadderFittedStepCount")
    control_req = _require_cols(control_cols, "IdentityKey", "Assay", "LadderQC", "LadderR2", "LadderExpectedStepCount", "LadderFittedStepCount")
    for row_idx, assay in enumerate(assays, start=2):
        ws.cell(row_idx, 1, assay)
        ws.cell(row_idx, 2, f'=COUNTIF({_range_ref("Patient_Runs", patient_req["Assay"])}, {assay_col})+COUNTIF({_range_ref("Control_Runs", control_req["Assay"])}, {assay_col})')
        ws.cell(row_idx, 3, f'=COUNTIF({_range_ref("Patient_Runs", patient_req["Assay"])}, {assay_col})')
        ws.cell(row_idx, 4, f'=COUNTIF({_range_ref("Control_Runs", control_req["Assay"])}, {assay_col})')
        ws.cell(row_idx, 5, (
            f'=COUNTIFS({_range_ref("Patient_Runs", patient_req["Assay"])}, {assay_col}, {_range_ref("Patient_Runs", patient_req["LadderQC"])}, "<>", {_range_ref("Patient_Runs", patient_req["LadderQC"])}, "<>ok")'
            f'+COUNTIFS({_range_ref("Control_Runs", control_req["Assay"])}, {assay_col}, {_range_ref("Control_Runs", control_req["LadderQC"])}, "<>", {_range_ref("Control_Runs", control_req["LadderQC"])}, "<>ok")'
        ))
        ws.cell(row_idx, 6, (
            f'=IFERROR((SUMIF({_range_ref("Patient_Runs", patient_req["Assay"])}, {assay_col}, {_range_ref("Patient_Runs", patient_req["LadderR2"])})'
            f'+SUMIF({_range_ref("Control_Runs", control_req["Assay"])}, {assay_col}, {_range_ref("Control_Runs", control_req["LadderR2"])}) )/$B{row_idx},0)'
        ))
        ws.cell(row_idx, 7, (
            f'=SUMPRODUCT(--({_range_ref("Patient_Runs", patient_req["Assay"], end_row=patient_last)}=$A{row_idx}),--({_range_ref("Patient_Runs", patient_req["LadderFittedStepCount"], end_row=patient_last)}<{_range_ref("Patient_Runs", patient_req["LadderExpectedStepCount"], end_row=patient_last)}))'
            f'+SUMPRODUCT(--({_range_ref("Control_Runs", control_req["Assay"], end_row=control_last)}=$A{row_idx}),--({_range_ref("Control_Runs", control_req["LadderFittedStepCount"], end_row=control_last)}<{_range_ref("Control_Runs", control_req["LadderExpectedStepCount"], end_row=control_last)}))'
        ))
        ws.cell(row_idx, 8, f'=IFERROR(E{row_idx}/B{row_idx},0)')
    _style_table(ws, 1, max(len(assays) + 1, 2), 1, len(headers))
    ws.freeze_panes = "A2"
    _autofit(ws)
    for cell in ws["F"][1:]:
        cell.number_format = "0.000000"
    for cell in ws["H"][1:]:
        cell.number_format = "0.0%"


def _build_run_summary(ws, assays: list[str], patient_cols: dict[str, str], patient_last: int) -> None:
    ws.sheet_properties.tabColor = "5B9BD5"
    headers = ["Assay", "Files", "ReviewFiles", "AvgR2", "PartialFits"]
    ws.append(headers)
    patient_req = _require_cols(patient_cols, "IdentityKey", "Assay", "LadderQC", "LadderR2", "LadderExpectedStepCount", "LadderFittedStepCount")
    for row_idx, assay in enumerate(assays, start=2):
        ws.cell(row_idx, 1, assay)
        ws.cell(row_idx, 2, f'=COUNTIF({_range_ref("Patient_Runs", patient_req["Assay"])},$A{row_idx})')
        ws.cell(row_idx, 3, f'=COUNTIFS({_range_ref("Patient_Runs", patient_req["Assay"])},$A{row_idx},{_range_ref("Patient_Runs", patient_req["LadderQC"])}, "<>", {_range_ref("Patient_Runs", patient_req["LadderQC"])}, "<>ok")')
        ws.cell(row_idx, 4, f'=IFERROR(SUMIF({_range_ref("Patient_Runs", patient_req["Assay"])},$A{row_idx},{_range_ref("Patient_Runs", patient_req["LadderR2"])})/B{row_idx},0)')
        ws.cell(row_idx, 5, f'=SUMPRODUCT(--({_range_ref("Patient_Runs", patient_req["Assay"], end_row=patient_last)}=$A{row_idx}),--({_range_ref("Patient_Runs", patient_req["LadderFittedStepCount"], end_row=patient_last)}<{_range_ref("Patient_Runs", patient_req["LadderExpectedStepCount"], end_row=patient_last)}))')
    _style_table(ws, 1, max(len(assays) + 1, 2), 1, len(headers))
    ws.freeze_panes = "A2"
    _autofit(ws)
    for cell in ws["D"][1:]:
        cell.number_format = "0.000000"


def _build_control_summary(ws, pairs: list[list[str]], control_cols: dict[str, str], control_last: int) -> None:
    ws.sheet_properties.tabColor = "5B9BD5"
    headers = ["Control", "Assay", "Files", "ReviewFiles", "AvgR2", "PartialFits"]
    ws.append(headers)
    control_req = _require_cols(control_cols, "IdentityKey", "Control", "Assay", "LadderQC", "LadderR2", "LadderExpectedStepCount", "LadderFittedStepCount")
    for row_idx, pair in enumerate(pairs, start=2):
        control_name, assay = pair
        ws.cell(row_idx, 1, control_name)
        ws.cell(row_idx, 2, assay)
        ws.cell(row_idx, 3, f'=COUNTIFS({_range_ref("Control_Runs", control_req["Control"])},$A{row_idx},{_range_ref("Control_Runs", control_req["Assay"])},$B{row_idx})')
        ws.cell(row_idx, 4, f'=COUNTIFS({_range_ref("Control_Runs", control_req["Control"])},$A{row_idx},{_range_ref("Control_Runs", control_req["Assay"])},$B{row_idx},{_range_ref("Control_Runs", control_req["LadderQC"])}, "<>", {_range_ref("Control_Runs", control_req["LadderQC"])}, "<>ok")')
        ws.cell(row_idx, 5, f'=IFERROR(SUMIFS({_range_ref("Control_Runs", control_req["LadderR2"])}, {_range_ref("Control_Runs", control_req["Control"])},$A{row_idx}, {_range_ref("Control_Runs", control_req["Assay"])},$B{row_idx})/C{row_idx},0)')
        ws.cell(row_idx, 6, f'=SUMPRODUCT(--({_range_ref("Control_Runs", control_req["Control"], end_row=control_last)}=$A{row_idx}),--({_range_ref("Control_Runs", control_req["Assay"], end_row=control_last)}=$B{row_idx}),--({_range_ref("Control_Runs", control_req["LadderFittedStepCount"], end_row=control_last)}<{_range_ref("Control_Runs", control_req["LadderExpectedStepCount"], end_row=control_last)}))')
    _style_table(ws, 1, max(len(pairs) + 1, 2), 1, len(headers))
    ws.freeze_panes = "A2"
    _autofit(ws)
    for cell in ws["E"][1:]:
        cell.number_format = "0.000000"


def _build_daily_trend(ws, run_days: list[str], patient_cols: dict[str, str], control_cols: dict[str, str]) -> None:
    ws.sheet_properties.tabColor = "70AD47"
    headers = ["RunDay", "Files", "ReviewFiles", "AvgR2"]
    ws.append(headers)
    patient_req = _require_cols(patient_cols, "IdentityKey", "RunDate", "LadderQC", "LadderR2")
    control_req = _require_cols(control_cols, "IdentityKey", "RunDate", "LadderQC", "LadderR2")
    for row_idx, run_day in enumerate(run_days, start=2):
        ws.cell(row_idx, 1, run_day)
        ws.cell(row_idx, 2, f'=COUNTIF({_range_ref("Patient_Runs", patient_req["RunDate"])},$A{row_idx})+COUNTIF({_range_ref("Control_Runs", control_req["RunDate"])},$A{row_idx})')
        ws.cell(row_idx, 3, (
            f'=COUNTIFS({_range_ref("Patient_Runs", patient_req["RunDate"])},$A{row_idx},{_range_ref("Patient_Runs", patient_req["LadderQC"])}, "<>", {_range_ref("Patient_Runs", patient_req["LadderQC"])}, "<>ok")'
            f'+COUNTIFS({_range_ref("Control_Runs", control_req["RunDate"])},$A{row_idx},{_range_ref("Control_Runs", control_req["LadderQC"])}, "<>", {_range_ref("Control_Runs", control_req["LadderQC"])}, "<>ok")'
        ))
        ws.cell(row_idx, 4, (
            f'=IFERROR((SUMIF({_range_ref("Patient_Runs", patient_req["RunDate"])},$A{row_idx},{_range_ref("Patient_Runs", patient_req["LadderR2"])})'
            f'+SUMIF({_range_ref("Control_Runs", control_req["RunDate"])},$A{row_idx},{_range_ref("Control_Runs", control_req["LadderR2"])}) )/B{row_idx},0)'
        ))
    _style_table(ws, 1, max(len(run_days) + 1, 2), 1, len(headers))
    ws.freeze_panes = "A2"
    _autofit(ws)
    for cell in ws["D"][1:]:
        cell.number_format = "0.000000"


def _build_pk_summary(ws, assays: list[str], pk_cols: dict[str, str], pk_last: int, *, kind: str, title: str) -> None:
    ws.sheet_properties.tabColor = "ED7D31" if kind == "sample" else "C55A11"
    headers = ["Assay", "MarkerRows", "MeanAbsDeltaBP", "MaxAbsDeltaBP", "Over2bp", "Over5bp", "AvgHeight"]
    ws.append(headers)
    pk_req = _require_cols(pk_cols, "IdentityKey", "Assay", "Kind", "AbsDeltaBP", "Height")
    for row_idx, assay in enumerate(assays, start=2):
        ws.cell(row_idx, 1, assay)
        ws.cell(row_idx, 2, f'=COUNTIFS({_range_ref("PK_Peaks", pk_req["Assay"])},$A{row_idx},{_range_ref("PK_Peaks", pk_req["Kind"])}, "{kind}")')
        ws.cell(row_idx, 3, f'=IFERROR(AVERAGEIFS({_range_ref("PK_Peaks", pk_req["AbsDeltaBP"])}, {_range_ref("PK_Peaks", pk_req["Assay"])},$A{row_idx}, {_range_ref("PK_Peaks", pk_req["Kind"])}, "{kind}"),0)')
        ws.cell(row_idx, 4, f'=IFERROR(MAXIFS({_range_ref("PK_Peaks", pk_req["AbsDeltaBP"])}, {_range_ref("PK_Peaks", pk_req["Assay"])},$A{row_idx}, {_range_ref("PK_Peaks", pk_req["Kind"])}, "{kind}"),0)')
        ws.cell(row_idx, 5, f'=SUMPRODUCT(--({_range_ref("PK_Peaks", pk_req["Assay"], end_row=pk_last)}=$A{row_idx}),--({_range_ref("PK_Peaks", pk_req["Kind"], end_row=pk_last)}="{kind}"),--({_range_ref("PK_Peaks", pk_req["AbsDeltaBP"], end_row=pk_last)}>2))')
        ws.cell(row_idx, 6, f'=SUMPRODUCT(--({_range_ref("PK_Peaks", pk_req["Assay"], end_row=pk_last)}=$A{row_idx}),--({_range_ref("PK_Peaks", pk_req["Kind"], end_row=pk_last)}="{kind}"),--({_range_ref("PK_Peaks", pk_req["AbsDeltaBP"], end_row=pk_last)}>5))')
        ws.cell(row_idx, 7, f'=IFERROR(AVERAGEIFS({_range_ref("PK_Peaks", pk_req["Height"])}, {_range_ref("PK_Peaks", pk_req["Assay"])},$A{row_idx}, {_range_ref("PK_Peaks", pk_req["Kind"])}, "{kind}"),0)')
    ws["A1"] = "Assay"
    _style_table(ws, 1, max(len(assays) + 1, 2), 1, len(headers))
    ws.freeze_panes = "A2"
    _autofit(ws)
    for col in ("C", "D"):
        for cell in ws[col][1:]:
            cell.number_format = "0.00"


def _write_static_frame(ws, title: str, df: pd.DataFrame) -> None:
    ws.sheet_properties.tabColor = "C00000"
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True)
    headers = list(df.columns)
    for idx, header in enumerate(headers, start=1):
        ws.cell(2, idx, header)
    for row_idx, row in enumerate(df.itertuples(index=False), start=3):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row_idx, col_idx, value)
    if headers:
        _style_table(ws, 2, max(len(df) + 2, 3), 1, len(headers))
    ws.freeze_panes = "A3"
    _autofit(ws)


def _add_card(ws, cell: str, label: str, formula: str, fill: PatternFill) -> None:
    ws[cell] = label
    ws[cell].fill = fill
    ws[cell].font = Font(size=10, bold=True, color="3A3A3A")
    ws[cell].alignment = Alignment(horizontal="center")
    ws[cell].border = BOX_BORDER
    value_cell = ws.cell(ws[cell].row + 1, ws[cell].column, formula)
    value_cell.fill = fill
    value_cell.font = Font(size=17, bold=True, color="1F1F1F")
    value_cell.alignment = Alignment(horizontal="center")
    value_cell.border = BOX_BORDER


def _build_dashboard(ws, assay_count: int, run_day_count: int, review_count: int, outlier_count: int, outlier_headers: list[str], assay_ws, control_ws, pk_sample_ws, pk_ladder_ws, daily_ws) -> None:
    ws.sheet_properties.tabColor = "2F75B5"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A9"
    for col, width in {
        "A": 18, "B": 16, "C": 16, "D": 16, "E": 18, "F": 18, "G": 18, "H": 18, "I": 16, "J": 16, "K": 16, "L": 16,
    }.items():
        ws.column_dimensions[col].width = width

    ws.merge_cells("A1:L2")
    ws["A1"] = "Fraggler Clonality Tracking Dashboard"
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A3"] = "This dashboard recalculates in Excel from the raw run sheets below."
    ws["A3"].font = BOLD
    ws["A4"] = "Formula note"
    ws["A4"].font = BOLD
    ws["B4"] = "Excel stores formulas in English internally; Norwegian Excel localizes them when opened."

    _add_card(ws, "A6", "Tracked Runs", '=COUNTIF(Patient_Runs!$A:$A,"<>")+COUNTIF(Control_Runs!$A:$A,"<>")', CARD_BLUE)
    _add_card(ws, "B6", "Unique Assays", f"={assay_count}", CARD_TEAL)
    patient_cols = _col_map(ws.parent["Patient_Runs"])
    control_cols = _col_map(ws.parent["Control_Runs"])
    pk_cols = _col_map(ws.parent["PK_Peaks"])
    _add_card(ws, "C6", "Review Files", f'=COUNTIFS(Patient_Runs!${patient_cols["LadderQC"]}:${patient_cols["LadderQC"]},"<>",Patient_Runs!${patient_cols["LadderQC"]}:${patient_cols["LadderQC"]},"<>ok")+COUNTIFS(Control_Runs!${control_cols["LadderQC"]}:${control_cols["LadderQC"]},"<>",Control_Runs!${control_cols["LadderQC"]}:${control_cols["LadderQC"]},"<>ok")', CARD_RED)
    _add_card(ws, "D6", "Ladder OK Rate", f'=IFERROR((COUNTIF(Patient_Runs!${patient_cols["LadderQC"]}:${patient_cols["LadderQC"]},"ok")+COUNTIF(Control_Runs!${control_cols["LadderQC"]}:${control_cols["LadderQC"]},"ok"))/(COUNTIF(Patient_Runs!$A:$A,"<>")+COUNTIF(Control_Runs!$A:$A,"<>")),0)', CARD_GOLD)
    _add_card(ws, "E6", "PK Marker Rows", f'=COUNTIF(PK_Peaks!${pk_cols["MarkerName"]}:${pk_cols["MarkerName"]},"<>")', CARD_BLUE)
    _add_card(ws, "F6", "PK Sample Mean |delta|", f'=IFERROR(AVERAGEIFS(PK_Peaks!${pk_cols["AbsDeltaBP"]}:${pk_cols["AbsDeltaBP"]},PK_Peaks!${pk_cols["Kind"]}:${pk_cols["Kind"]},"sample"),0)', CARD_ORANGE)
    _add_card(ws, "G6", "PK Sample >2 bp", f'=COUNTIFS(PK_Peaks!${pk_cols["Kind"]}:${pk_cols["Kind"]},"sample",PK_Peaks!${pk_cols["AbsDeltaBP"]}:${pk_cols["AbsDeltaBP"]},">2")', CARD_RED)
    _add_card(ws, "H6", "PK Ladder Mean |delta|", f'=IFERROR(AVERAGEIFS(PK_Peaks!${pk_cols["AbsDeltaBP"]}:${pk_cols["AbsDeltaBP"]},PK_Peaks!${pk_cols["Kind"]}:${pk_cols["Kind"]},"ladder"),0)', CARD_TEAL)
    ws["D7"].number_format = "0.0%"
    ws["F7"].number_format = '0.00 "bp"'
    ws["H7"].number_format = '0.00 "bp"'

    ws["A9"] = "Ladder Overview"
    ws["A9"].font = Font(size=13, bold=True)
    overview_headers = ["Scope", "Runs", "Ladder OK", "Review Required", "OK Rate", "Avg R2", "Partial Fits"]
    for idx, header in enumerate(overview_headers, start=1):
        ws.cell(10, idx, header)
    patient_req = _require_cols(patient_cols, "IdentityKey", "Assay", "LadderQC", "LadderR2", "LadderExpectedStepCount", "LadderFittedStepCount")
    control_req = _require_cols(control_cols, "IdentityKey", "Assay", "LadderQC", "LadderR2", "LadderExpectedStepCount", "LadderFittedStepCount")
    overview_rows = [
        ("All", '=COUNTIF(Patient_Runs!$A:$A,"<>")+COUNTIF(Control_Runs!$A:$A,"<>")', f'=COUNTIF({_range_ref("Patient_Runs", patient_req["LadderQC"])}, "ok")+COUNTIF({_range_ref("Control_Runs", control_req["LadderQC"])}, "ok")', f'=COUNTIFS({_range_ref("Patient_Runs", patient_req["LadderQC"])}, "<>", {_range_ref("Patient_Runs", patient_req["LadderQC"])}, "<>ok")+COUNTIFS({_range_ref("Control_Runs", control_req["LadderQC"])}, "<>", {_range_ref("Control_Runs", control_req["LadderQC"])}, "<>ok")', '=IFERROR(C11/B11,0)', f'=IFERROR((SUM({_range_ref("Patient_Runs", patient_req["LadderR2"])} )+SUM({_range_ref("Control_Runs", control_req["LadderR2"])}))/B11,0)', f'=SUMPRODUCT(--({_range_ref("Patient_Runs", patient_req["LadderFittedStepCount"], end_row=1048576)}<{_range_ref("Patient_Runs", patient_req["LadderExpectedStepCount"], end_row=1048576)}))+SUMPRODUCT(--({_range_ref("Control_Runs", control_req["LadderFittedStepCount"], end_row=1048576)}<{_range_ref("Control_Runs", control_req["LadderExpectedStepCount"], end_row=1048576)}))'),
        ("Patient", '=COUNTIF(Patient_Runs!$A:$A,"<>")', f'=COUNTIF({_range_ref("Patient_Runs", patient_req["LadderQC"])}, "ok")', f'=COUNTIFS({_range_ref("Patient_Runs", patient_req["LadderQC"])}, "<>", {_range_ref("Patient_Runs", patient_req["LadderQC"])}, "<>ok")', '=IFERROR(C12/B12,0)', f'=IFERROR(SUM({_range_ref("Patient_Runs", patient_req["LadderR2"])})/B12,0)', f'=SUMPRODUCT(--({_range_ref("Patient_Runs", patient_req["LadderFittedStepCount"], end_row=1048576)}<{_range_ref("Patient_Runs", patient_req["LadderExpectedStepCount"], end_row=1048576)}))'),
        ("Control", '=COUNTIF(Control_Runs!$A:$A,"<>")', f'=COUNTIF({_range_ref("Control_Runs", control_req["LadderQC"])}, "ok")', f'=COUNTIFS({_range_ref("Control_Runs", control_req["LadderQC"])}, "<>", {_range_ref("Control_Runs", control_req["LadderQC"])}, "<>ok")', '=IFERROR(C13/B13,0)', f'=IFERROR(SUM({_range_ref("Control_Runs", control_req["LadderR2"])})/B13,0)', f'=SUMPRODUCT(--({_range_ref("Control_Runs", control_req["LadderFittedStepCount"], end_row=1048576)}<{_range_ref("Control_Runs", control_req["LadderExpectedStepCount"], end_row=1048576)}))'),
    ]
    for row_idx, row in enumerate(overview_rows, start=11):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row_idx, col_idx, value)
    _style_table(ws, 10, 13, 1, 7)
    for col in ("E", "F"):
        for cell in ws[col][10:13]:
            cell.number_format = "0.000000" if col == "F" else "0.0%"

    ws["J9"] = "PK Sample Delta Focus"
    ws["J9"].font = Font(size=13, bold=True)
    focus_headers = ["Assay", "MarkerRows", "MeanAbsDeltaBP", "MaxAbsDeltaBP", "Over2bp", "Over5bp", "AvgHeight"]
    for idx, header in enumerate(focus_headers, start=10):
        ws.cell(10, idx, header)
    for offset in range(1, min(assay_count, 8) + 1):
        src = offset + 1
        dst = 10 + offset
        for col_idx, col_letter in enumerate(("A", "B", "C", "D", "E", "F", "G"), start=10):
            ws.cell(dst, col_idx, f"='PK_Sample_Delta'!{col_letter}{src}")
    _style_table(ws, 10, 10 + max(min(assay_count, 8), 1), 10, 16)
    for col in ("L", "M"):
        for cell in ws[col][10:10 + max(min(assay_count, 8), 1)]:
            cell.number_format = "0.00"

    ws["A16"] = "Assay Watchlist"
    ws["A16"].font = Font(size=13, bold=True)
    watch_headers = ["Assay", "Files", "PatientFiles", "ControlFiles", "LadderReview", "AvgR2", "PartialFits", "ReviewRate"]
    for idx, header in enumerate(watch_headers, start=1):
        ws.cell(17, idx, header)
    for offset in range(1, min(assay_count, 8) + 1):
        src = offset + 1
        dst = 17 + offset
        for col_idx, col_letter in enumerate(("A", "B", "C", "D", "E", "F", "G", "H"), start=1):
            ws.cell(dst, col_idx, f"='Assay_Summary'!{col_letter}{src}")
    _style_table(ws, 17, 17 + max(min(assay_count, 8), 1), 1, 8)
    for cell in ws["F"][17:17 + max(min(assay_count, 8), 1)]:
        cell.number_format = "0.000000"
    for cell in ws["H"][17:17 + max(min(assay_count, 8), 1)]:
        cell.number_format = "0.0%"

    ws["J16"] = "Top PK Outliers"
    ws["J16"].font = Font(size=13, bold=True)
    for idx, header in enumerate(outlier_headers, start=10):
        ws.cell(17, idx, header)
    for offset in range(1, min(outlier_count, 8) + 1):
        src = offset + 2
        dst = 17 + offset
        for col_idx, col_letter in enumerate(tuple(get_column_letter(i) for i in range(1, len(outlier_headers) + 1)), start=10):
            ws.cell(dst, col_idx, f"='PK_Outliers'!{col_letter}{src}")
    _style_table(ws, 17, 17 + max(min(outlier_count, 8), 1), 10, 19)

    ws.conditional_formatting.add(
        f"E18:E{17 + max(min(assay_count, 8), 1)}",
        ColorScaleRule(start_type="num", start_value=0, start_color="E2F0D9", mid_type="percentile", mid_value=50, mid_color="FFE699", end_type="max", end_color="F4CCCC"),
    )
    ws.conditional_formatting.add(
        f"L11:L{10 + max(min(assay_count, 8), 1)}",
        ColorScaleRule(start_type="num", start_value=0, start_color="E2F0D9", mid_type="percentile", mid_value=50, mid_color="FFE699", end_type="max", end_color="F4CCCC"),
    )

    status_chart = BarChart()
    status_chart.title = "Ladder QC Status"
    status_chart.y_axis.title = "Runs"
    status_chart.height = 7
    status_chart.width = 9
    status_chart.add_data(Reference(ws, min_col=3, min_row=10, max_row=13), titles_from_data=True)
    status_chart.set_categories(Reference(ws, min_col=1, min_row=11, max_row=13))
    ws.add_chart(status_chart, "A28")

    assay_chart = BarChart()
    assay_chart.title = "Files by Assay"
    assay_chart.y_axis.title = "Files"
    assay_chart.height = 7
    assay_chart.width = 10
    assay_chart.add_data(Reference(assay_ws, min_col=2, min_row=1, max_row=min(11, assay_count + 1)), titles_from_data=True)
    assay_chart.set_categories(Reference(assay_ws, min_col=1, min_row=2, max_row=min(11, assay_count + 1)))
    ws.add_chart(assay_chart, "F28")

    delta_chart = BarChart()
    delta_chart.title = "PK Sample Mean |delta bp| by Assay"
    delta_chart.y_axis.title = "|Delta bp|"
    delta_chart.height = 7
    delta_chart.width = 10
    delta_chart.add_data(Reference(pk_sample_ws, min_col=3, min_row=1, max_row=min(11, assay_count + 1)), titles_from_data=True)
    delta_chart.set_categories(Reference(pk_sample_ws, min_col=1, min_row=2, max_row=min(11, assay_count + 1)))
    ws.add_chart(delta_chart, "A44")

    daily_chart = LineChart()
    daily_chart.title = "Daily Average Ladder R2"
    daily_chart.y_axis.title = "Average R2"
    daily_chart.height = 7
    daily_chart.width = 10
    daily_chart.add_data(Reference(daily_ws, min_col=4, min_row=1, max_row=max(run_day_count + 1, 2)), titles_from_data=True)
    daily_chart.set_categories(Reference(daily_ws, min_col=1, min_row=2, max_row=max(run_day_count + 1, 2)))
    ws.add_chart(daily_chart, "F44")
