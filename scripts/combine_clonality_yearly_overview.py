from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


TARGET_SHEETS = ("Patient_Runs", "Control_Runs", "PK_Peaks")
WEAK_LADDER_STRATEGIES = {"auto_partial", "high_end_rescue"}


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Combine monthly clonality tracking workbooks into one yearly overview workbook."
    )
    parser.add_argument(
        "--run-root",
        type=Path,
        required=True,
        help="Full run root containing month_runs/<YEAR_MM>/track-clonality.xlsx",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Output workbook path. Defaults to <run-root>/track-clonality-<YEAR>-overview.xlsx",
    )
    parser.add_argument(
        "--year-label",
        default="2025",
        help="Label used in output sheet names, for example 2025 or 2024.",
    )
    return parser


def _read_month_sheet(workbook_path: Path, month: str, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(workbook_path, sheet_name=sheet_name)
    if "Month" in df.columns:
        df["Month"] = month
        cols = ["Month"] + [column for column in df.columns if column != "Month"]
        df = df[cols]
    else:
        df.insert(0, "Month", month)
    return df


def _auto_fit_columns(ws) -> None:
    for column_cells in ws.columns:
        lengths = []
        for cell in column_cells:
            if cell.value is None:
                continue
            lengths.append(len(str(cell.value)))
        if not lengths:
            continue
        width = min(max(lengths) + 2, 40)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def _style_workbook(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
        ws.freeze_panes = "A2"
        _auto_fit_columns(ws)

    if "Overview" in wb.sheetnames:
        ws = wb["Overview"]
        ws.freeze_panes = "A1"
        if ws["A1"].value:
            ws["A1"].font = Font(bold=True, size=14)

    wb.save(workbook_path)


def combine_run_root(run_root: Path, output_path: Path, *, year_label: str = "2025") -> Path:
    month_runs_root = run_root / "month_runs"
    if not month_runs_root.is_dir():
        raise FileNotFoundError(f"Missing month_runs directory: {month_runs_root}")

    month_names = sorted(p.name for p in month_runs_root.iterdir() if p.is_dir())
    combined: dict[str, list[pd.DataFrame]] = {sheet: [] for sheet in TARGET_SHEETS}
    overview_rows: list[dict[str, object]] = []

    for month in month_names:
        workbook_path = month_runs_root / month / "track-clonality.xlsx"
        if not workbook_path.exists():
            continue
        xl = pd.ExcelFile(workbook_path)
        month_summary: dict[str, object] = {"Month": month}
        for sheet_name in TARGET_SHEETS:
            if sheet_name not in xl.sheet_names:
                month_summary[f"{sheet_name}_Rows"] = 0
                continue
            df = _read_month_sheet(workbook_path, month, sheet_name)
            combined[sheet_name].append(df)
            month_summary[f"{sheet_name}_Rows"] = int(len(df))
            if sheet_name == "Patient_Runs" and not df.empty:
                month_summary["Patient_LadderReviewRequired"] = int((df["LadderQC"] == "review_required").sum())
                month_summary["Patient_AutoPartial"] = int((df["LadderFitStrategy"] == "auto_partial").sum())
            if sheet_name == "PK_Peaks" and not df.empty:
                month_summary["PK_Outliers_AbsDeltaGT2"] = int((pd.to_numeric(df["AbsDeltaBP"], errors="coerce") > 2.0).sum())
        overview_rows.append(month_summary)

    overview_df = pd.DataFrame(overview_rows).fillna(0)
    patient_df = pd.concat(combined["Patient_Runs"], ignore_index=True) if combined["Patient_Runs"] else pd.DataFrame()
    control_df = pd.concat(combined["Control_Runs"], ignore_index=True) if combined["Control_Runs"] else pd.DataFrame()
    pk_df = pd.concat(combined["PK_Peaks"], ignore_index=True) if combined["PK_Peaks"] else pd.DataFrame()

    weak_ladders_df = pd.DataFrame()
    if not patient_df.empty:
        weak_mask = (
            (patient_df.get("LadderQC", pd.Series(dtype=object)) == "review_required")
            | patient_df.get("LadderFitStrategy", pd.Series(dtype=object)).isin(WEAK_LADDER_STRATEGIES)
        )
        weak_ladders_df = patient_df.loc[weak_mask].copy()
        if not weak_ladders_df.empty:
            preferred_columns = [
                "Month",
                "SourceRunDir",
                "IdentityKey",
                "Assay",
                "Well",
                "RunDate",
                "RunCode",
                "Ladder",
                "LadderQC",
                "LadderFitStrategy",
                "LadderFittedStepCount",
                "LadderExpectedStepCount",
                "LadderR2",
            ]
            weak_ladders_df = weak_ladders_df[[column for column in preferred_columns if column in weak_ladders_df.columns]]

    pk_outliers_df = pd.DataFrame()
    if not pk_df.empty and "AbsDeltaBP" in pk_df.columns:
        abs_delta = pd.to_numeric(pk_df["AbsDeltaBP"], errors="coerce").abs()
        pk_outliers_df = pk_df.loc[abs_delta > 2.0].copy()
        if not pk_outliers_df.empty:
            preferred_columns = [
                "Month",
                "SourceRunDir",
                "IdentityKey",
                "Assay",
                "Marker",
                "SampleKind",
                "RunDate",
                "RunCode",
                "Well",
                "ExpectedBP",
                "ObservedBP",
                "DeltaBP",
                "AbsDeltaBP",
                "PeakHeight",
                "PeakTime",
            ]
            pk_outliers_df = pk_outliers_df[[column for column in preferred_columns if column in pk_outliers_df.columns]]

    total_row = {
        "Month": "TOTAL",
        "Patient_Runs_Rows": int(len(patient_df)),
        "Control_Runs_Rows": int(len(control_df)),
        "PK_Peaks_Rows": int(len(pk_df)),
        "Patient_LadderReviewRequired": int((patient_df.get("LadderQC", pd.Series(dtype=object)) == "review_required").sum()) if not patient_df.empty else 0,
        "Patient_AutoPartial": int((patient_df.get("LadderFitStrategy", pd.Series(dtype=object)) == "auto_partial").sum()) if not patient_df.empty else 0,
        "PK_Outliers_AbsDeltaGT2": int((pd.to_numeric(pk_df.get("AbsDeltaBP", pd.Series(dtype=float)), errors="coerce") > 2.0).sum()) if not pk_df.empty else 0,
    }
    overview_df = pd.concat([overview_df, pd.DataFrame([total_row])], ignore_index=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        overview_df.to_excel(writer, sheet_name="Overview", index=False)
        patient_df.to_excel(writer, sheet_name=f"Patient_Runs_{year_label}", index=False)
        control_df.to_excel(writer, sheet_name=f"Control_Runs_{year_label}", index=False)
        pk_df.to_excel(writer, sheet_name=f"PK_Peaks_{year_label}", index=False)
        weak_ladders_df.to_excel(writer, sheet_name=f"Weak_Ladders_{year_label}", index=False)
        pk_outliers_df.to_excel(writer, sheet_name=f"PK_Outliers_{year_label}", index=False)

    _style_workbook(output_path)
    return output_path


def main(argv: list[str] | None = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)
    output_path = args.output or (args.run_root / f"track-clonality-{args.year_label}-overview.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    result = combine_run_root(args.run_root, output_path, year_label=str(args.year_label))
    print(result)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
