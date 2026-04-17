from __future__ import annotations

import argparse
import csv
import json
import shutil
import sys
from collections import defaultdict
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import APP_SETTINGS
from core.analyses.flt3.qc_tracker import FLT3_NPM1_QC_TRACKER_FILENAME
from core.runner import run_pipeline_job_collect

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))


from scripts.validate_flt3_ladder_fits import (  # noqa: E402
    DEFAULT_DATA_ROOT,
    DEFAULT_OUTPUT_ROOT,
    run_validation as run_ladder_validation,
)


DEFAULT_REQUIRED_RUN_NAME = "3730DNA"
DEFAULT_EXCLUDED_BASENAMES = [
    "25OUM11314_p2_RATIO__310725_F04_H9C0ZIZJ.fsa",
    "25OUM12881_itd-Ratio__250825_F04_C990WO66.fsa",
    "26OUM01055_ITD_10x_27012026_E01_H9C0VCGS.fsa",
    "25OUM12253_RATIO__130825_A03_C990WO65.fsa",
]


def _timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Run a FLT3 ladder-fit backfill validation across a data tree, filtered to a "
            "specific run-name signature such as 3730DNA, and export residual summaries."
        )
    )
    parser.add_argument(
        "--data-root",
        type=Path,
        default=DEFAULT_DATA_ROOT,
        help=f"Root containing FLT3 year folders. Default: {DEFAULT_DATA_ROOT}",
    )
    parser.add_argument(
        "--output-root",
        type=Path,
        default=DEFAULT_OUTPUT_ROOT,
        help=f"Root directory for validation outputs. Default: {DEFAULT_OUTPUT_ROOT}",
    )
    parser.add_argument(
        "--run-name",
        type=str,
        default=None,
        help="Optional run directory name. Defaults to flt3_backfill_<timestamp>.",
    )
    parser.add_argument(
        "--year",
        dest="years",
        action="append",
        default=[],
        help="Optional year folder to include. Repeat for multiple years.",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=6,
        help="Worker count passed through to validate_flt3_ladder_fits.py.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Optional file limit after filtering. 0 means no limit.",
    )
    parser.add_argument(
        "--input-manifest",
        type=Path,
        default=None,
        help="Optional CSV manifest with a path column.",
    )
    parser.add_argument(
        "--include-npm1",
        action="store_true",
        help="Include NPM1 rows alongside FLT3.",
    )
    parser.add_argument(
        "--dit-only",
        action="store_true",
        help="Restrict to DIT specimen ids matching NNOU MNNNNN.",
    )
    parser.add_argument(
        "--timeout-seconds",
        type=int,
        default=45,
        help="Per-file timeout passed through to the FLT3 validator.",
    )
    parser.add_argument(
        "--checkpoint-every",
        type=int,
        default=100,
        help="Checkpoint cadence for partial output bundles.",
    )
    parser.add_argument(
        "--require-run-name-contains",
        type=str,
        default=DEFAULT_REQUIRED_RUN_NAME,
        help=f"Only validate files whose run_name contains this substring. Default: {DEFAULT_REQUIRED_RUN_NAME}",
    )
    parser.add_argument(
        "--exclude-basename",
        dest="excluded_basenames",
        action="append",
        default=[],
        help="Basename to exclude from archive validation. Repeat for multiple known human/machine-error files.",
    )
    return parser


def _coerce_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    if np.isfinite(result):
        return result
    return None


def _load_metrics_rows(metrics_csv: Path) -> list[dict[str, Any]]:
    with metrics_csv.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def _numeric_residuals(rows: list[dict[str, Any]], field: str) -> list[float]:
    values: list[float] = []
    for row in rows:
        value = _coerce_float(row.get(field))
        if value is not None:
            values.append(value)
    return values


def _residual_stats(rows: list[dict[str, Any]]) -> dict[str, Any]:
    max_vals = _numeric_residuals(rows, "max_abs_error_bp")
    mean_vals = _numeric_residuals(rows, "mean_abs_error_bp")
    r2_vals = _numeric_residuals(rows, "ladder_r2")

    def _quantiles(values: list[float]) -> dict[str, float] | None:
        if not values:
            return None
        arr = np.asarray(values, dtype=float)
        return {
            "median": float(np.median(arr)),
            "p90": float(np.quantile(arr, 0.90)),
            "p95": float(np.quantile(arr, 0.95)),
            "max": float(np.max(arr)),
        }

    return {
        "row_count": len(rows),
        "max_abs_error_bp": _quantiles(max_vals),
        "mean_abs_error_bp": _quantiles(mean_vals),
        "ladder_r2": _quantiles(r2_vals),
    }


def _group_residual_rows(rows: list[dict[str, Any]], field: str) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        key = str(row.get(field) or "")
        grouped.setdefault(key, []).append(row)

    exported: list[dict[str, Any]] = []
    for key, items in sorted(grouped.items(), key=lambda item: item[0]):
        stats = _residual_stats(items)
        exported.append(
            {
                field: key,
                "row_count": stats["row_count"],
                "median_max_abs_error_bp": (stats["max_abs_error_bp"] or {}).get("median", ""),
                "p90_max_abs_error_bp": (stats["max_abs_error_bp"] or {}).get("p90", ""),
                "p95_max_abs_error_bp": (stats["max_abs_error_bp"] or {}).get("p95", ""),
                "max_abs_error_bp": (stats["max_abs_error_bp"] or {}).get("max", ""),
                "median_mean_abs_error_bp": (stats["mean_abs_error_bp"] or {}).get("median", ""),
                "median_ladder_r2": (stats["ladder_r2"] or {}).get("median", ""),
                "status_counts": json.dumps(dict(Counter(str(r.get("status") or "") for r in items)), sort_keys=True),
                "strategy_counts": json.dumps(dict(Counter(str(r.get("ladder_fit_strategy") or "") for r in items)), sort_keys=True),
            }
        )
    return exported


def _worst_residual_rows(rows: list[dict[str, Any]], limit: int = 100) -> list[dict[str, Any]]:
    scored: list[tuple[float, dict[str, Any]]] = []
    for row in rows:
        max_abs = _coerce_float(row.get("max_abs_error_bp"))
        if max_abs is None:
            continue
        scored.append((max_abs, row))
    scored.sort(key=lambda item: (-item[0], str(item[1].get("path") or "")))
    exported: list[dict[str, Any]] = []
    for rank, (_, row) in enumerate(scored[:limit], start=1):
        exported.append(
            {
                "rank": rank,
                "path": row.get("path", ""),
                "relative_path": row.get("relative_path", ""),
                "file": row.get("file", ""),
                "year": row.get("year", ""),
                "run_name": row.get("run_name", ""),
                "assay": row.get("assay", ""),
                "analysis_type": row.get("analysis_type", ""),
                "status": row.get("status", ""),
                "ladder_fit_strategy": row.get("ladder_fit_strategy", ""),
                "review_reason": row.get("review_reason", ""),
                "max_abs_error_bp": row.get("max_abs_error_bp", ""),
                "mean_abs_error_bp": row.get("mean_abs_error_bp", ""),
                "ladder_r2": row.get("ladder_r2", ""),
            }
        )
    return exported


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in fieldnames:
                fieldnames.append(key)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _style_workbook(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    overview_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows():
            for cell in row:
                if ws.title == "Overview" and cell.row > 1 and cell.column == 1:
                    cell.fill = overview_fill
                    cell.font = Font(bold=True)
                if cell.row > 1 and isinstance(cell.value, (int, float)) and ws.title != "Overview":
                    cell.alignment = Alignment(horizontal="right")
        for idx, column_cells in enumerate(ws.columns, start=1):
            sample = [str(cell.value) if cell.value is not None else "" for cell in column_cells[:200]]
            max_len = max((len(v) for v in sample), default=0)
            ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 48)
    wb.save(workbook_path)


def build_excel_workbook(run_dir: Path, residual_summary: dict[str, Any], validator_summary: dict[str, Any]) -> Path:
    workbook_path = run_dir / "track-flt3-archive.xlsx"

    metrics = pd.read_csv(run_dir / "flt3_ladder_metrics.csv").fillna("")
    review = pd.read_csv(run_dir / "flt3_ladder_review_manifest.csv").fillna("")
    fit_failed = pd.read_csv(run_dir / "flt3_ladder_fit_failed_manifest.csv").fillna("")
    manual_adj = pd.read_csv(run_dir / "flt3_ladder_manual_adjustments_applied.csv").fillna("")
    residual_by_year = pd.read_csv(run_dir / "residual_by_year.csv").fillna("")
    residual_by_assay = pd.read_csv(run_dir / "residual_by_assay.csv").fillna("")
    residual_by_analysis = pd.read_csv(run_dir / "residual_by_analysis_type.csv").fillna("")
    residual_by_strategy = pd.read_csv(run_dir / "residual_by_strategy.csv").fillna("")
    residual_worst = pd.read_csv(run_dir / "residual_worst_cases.csv").fillna("")

    overview = pd.DataFrame(
        [
            {"Metric": "Generated UTC", "Value": residual_summary.get("generated_at_utc", "")},
            {"Metric": "Run Directory", "Value": str(run_dir)},
            {"Metric": "Validated Rows", "Value": int(residual_summary.get("validated_row_count", 0) or 0)},
            {"Metric": "Manual Review Count", "Value": int(validator_summary.get("manual_review_count", 0) or 0)},
            {
                "Metric": "Median Max Residual (bp)",
                "Value": ((residual_summary.get("overall") or {}).get("max_abs_error_bp") or {}).get("median", ""),
            },
            {
                "Metric": "P95 Max Residual (bp)",
                "Value": ((residual_summary.get("overall") or {}).get("max_abs_error_bp") or {}).get("p95", ""),
            },
            {
                "Metric": "Worst Max Residual (bp)",
                "Value": ((residual_summary.get("overall") or {}).get("max_abs_error_bp") or {}).get("max", ""),
            },
            {
                "Metric": "Median Ladder R2",
                "Value": ((residual_summary.get("overall") or {}).get("ladder_r2") or {}).get("median", ""),
            },
            {"Metric": "Excluded Files", "Value": int(validator_summary.get("excluded_count", 0) or 0)},
            {
                "Metric": "Excluded Basenames",
                "Value": ", ".join(validator_summary.get("excluded_basenames", []) or []),
            },
            {
                "Metric": "Status Counts",
                "Value": json.dumps(validator_summary.get("status_counts", {}), ensure_ascii=False, sort_keys=True),
            },
            {
                "Metric": "Queue Group Counts",
                "Value": json.dumps(validator_summary.get("queue_group_counts", {}), ensure_ascii=False, sort_keys=True),
            },
            {
                "Metric": "Strategy Counts",
                "Value": json.dumps(validator_summary.get("strategy_counts", {}), ensure_ascii=False, sort_keys=True),
            },
        ]
    )

    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        overview.to_excel(writer, sheet_name="Overview", index=False)
        residual_by_year.to_excel(writer, sheet_name="Residual_By_Year", index=False)
        residual_by_assay.to_excel(writer, sheet_name="Residual_By_Assay", index=False)
        residual_by_analysis.to_excel(writer, sheet_name="Residual_By_Analysis", index=False)
        residual_by_strategy.to_excel(writer, sheet_name="Residual_By_Strategy", index=False)
        review.to_excel(writer, sheet_name="Review_Queue", index=False)
        fit_failed.to_excel(writer, sheet_name="Fit_Failed", index=False)
        manual_adj.to_excel(writer, sheet_name="Manual_Adjustments", index=False)
        residual_worst.to_excel(writer, sheet_name="Worst_Residuals", index=False)
        metrics.to_excel(writer, sheet_name="All_Metrics", index=False)

    _style_workbook(workbook_path)
    return workbook_path


def _archive_metric_paths(metrics_csv: Path) -> list[Path]:
    if not metrics_csv.exists():
        return []

    discovered: list[Path] = []
    seen: set[Path] = set()
    with metrics_csv.open(newline="", encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            raw_path = str(row.get("path") or "").strip()
            if not raw_path:
                continue
            candidate = Path(raw_path).expanduser()
            if not candidate.exists() or candidate in seen:
                continue
            seen.add(candidate)
            discovered.append(candidate)
    return discovered


def build_tracking_workbook(run_dir: Path, *, status_callback=None) -> Path | None:
    metrics_csv = run_dir / "flt3_ladder_metrics.csv"
    tracker_path = run_dir / FLT3_NPM1_QC_TRACKER_FILENAME
    selected_files = _archive_metric_paths(metrics_csv)
    if not selected_files:
        return None

    grouped_files: dict[Path, list[Path]] = defaultdict(list)
    for path in selected_files:
        grouped_files[path.parent].append(path)

    tracker_tmp_root = run_dir / "_tracker_build_tmp"
    if tracker_path.exists():
        tracker_path.unlink()
    if tracker_tmp_root.exists():
        shutil.rmtree(tracker_tmp_root, ignore_errors=True)
    tracker_tmp_root.mkdir(parents=True, exist_ok=True)

    active_analysis_backup = APP_SETTINGS.get("active_analysis", "")
    APP_SETTINGS["active_analysis"] = "flt3"
    try:
        folder_total = len(grouped_files)
        for idx, folder in enumerate(sorted(grouped_files), start=1):
            if status_callback is not None:
                message = f"Building FLT3 tracker workbook ({idx}/{folder_total} folders)"
                try:
                    status_callback.emit(message)
                except AttributeError:
                    status_callback(message)
            run_pipeline_job_collect(
                fsa_dir=folder,
                base_outdir=tracker_tmp_root,
                out_folder_name="TRACKER_BUILD",
                scope="all",
                needle="",
                files=sorted(grouped_files[folder]),
                chunk_files=False,
                tracking_excel_path=tracker_path,
                update_tracking_workbook=True,
            )
    finally:
        APP_SETTINGS["active_analysis"] = active_analysis_backup
        shutil.rmtree(tracker_tmp_root, ignore_errors=True)

    return tracker_path if tracker_path.exists() else None


def build_residual_bundle(run_dir: Path) -> dict[str, Any]:
    metrics_csv = run_dir / "flt3_ladder_metrics.csv"
    rows = _load_metrics_rows(metrics_csv)
    validated_rows = [row for row in rows if not str(row.get("status") or "").startswith("skipped")]

    residual_summary = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "metrics_csv": str(metrics_csv),
        "validated_row_count": len(validated_rows),
        "overall": _residual_stats(validated_rows),
    }

    residual_summary_path = run_dir / "residual_summary.json"
    residual_summary_path.write_text(json.dumps(residual_summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")

    by_year = _group_residual_rows(validated_rows, "year")
    by_assay = _group_residual_rows(validated_rows, "assay")
    by_analysis_type = _group_residual_rows(validated_rows, "analysis_type")
    by_strategy = _group_residual_rows(validated_rows, "ladder_fit_strategy")
    worst_rows = _worst_residual_rows(validated_rows, limit=100)

    _write_csv(run_dir / "residual_by_year.csv", by_year)
    _write_csv(run_dir / "residual_by_assay.csv", by_assay)
    _write_csv(run_dir / "residual_by_analysis_type.csv", by_analysis_type)
    _write_csv(run_dir / "residual_by_strategy.csv", by_strategy)
    _write_csv(run_dir / "residual_worst_cases.csv", worst_rows)

    summary_md = f"""# FLT3 3730 Backfill Residual Summary

Generated UTC: {residual_summary["generated_at_utc"]}

Validated rows: **{residual_summary["validated_row_count"]}**
Median max residual bp: **{((residual_summary["overall"].get("max_abs_error_bp") or {}).get("median"))}**
P95 max residual bp: **{((residual_summary["overall"].get("max_abs_error_bp") or {}).get("p95"))}**
Worst max residual bp: **{((residual_summary["overall"].get("max_abs_error_bp") or {}).get("max"))}**
Median ladder R2: **{((residual_summary["overall"].get("ladder_r2") or {}).get("median"))}**

Files written:
- `residual_summary.json`
- `residual_by_year.csv`
- `residual_by_assay.csv`
- `residual_by_analysis_type.csv`
- `residual_by_strategy.csv`
- `residual_worst_cases.csv`
"""
    (run_dir / "residual_summary.md").write_text(summary_md, encoding="utf-8")
    return residual_summary


def run_backfill_validation(
    *,
    data_root: Path,
    output_root: Path,
    run_name: str | None,
    years: list[str],
    workers: int,
    limit: int,
    input_manifest: Path | None,
    include_npm1: bool,
    dit_only: bool,
    timeout_seconds: int,
    checkpoint_every: int,
    required_run_name_contains: str,
    excluded_basenames: list[str] | None = None,
    progress_callback=None,
    progress_max_callback=None,
    status_callback=None,
) -> dict[str, Any]:
    output_root = output_root.expanduser().resolve()
    output_root.mkdir(parents=True, exist_ok=True)
    safe_run_name = run_name or f"flt3_backfill_{_timestamp()}"
    run_dir = output_root / safe_run_name
    run_dir.mkdir(parents=True, exist_ok=True)

    summary = run_ladder_validation(
        data_root.expanduser().resolve(),
        run_dir,
        workers=max(1, int(workers)),
        selected_years=[str(year) for year in years],
        limit=max(0, int(limit)),
        input_manifest=input_manifest.expanduser().resolve() if input_manifest else None,
        include_npm1=bool(include_npm1),
        suppress_worker_output=True,
        timeout_seconds=max(0, int(timeout_seconds)),
        checkpoint_every=max(0, int(checkpoint_every)),
        dit_only=bool(dit_only),
        required_run_name=str(required_run_name_contains or ""),
        excluded_basenames=list(excluded_basenames or []),
        progress_callback=progress_callback,
        progress_max_callback=progress_max_callback,
        status_callback=status_callback,
    )
    (run_dir / "summary.json").write_text(
        json.dumps(summary, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )

    residual_summary = build_residual_bundle(run_dir)
    residual_workbook_path = build_excel_workbook(run_dir, residual_summary, summary)
    tracking_workbook_path = build_tracking_workbook(run_dir, status_callback=status_callback)
    payload = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "run_dir": str(run_dir),
        "summary_json": str(run_dir / "summary.json"),
        "residual_summary_json": str(run_dir / "residual_summary.json"),
        "workbook_path": str(tracking_workbook_path or residual_workbook_path),
        "tracking_workbook_path": str(tracking_workbook_path) if tracking_workbook_path else "",
        "residual_workbook_path": str(residual_workbook_path),
        "validator_summary": summary,
        "residual_summary": residual_summary,
    }
    (run_dir / "backfill_summary.json").write_text(
        json.dumps(payload, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    return payload


def main(argv: list[str] | None = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)
    payload = run_backfill_validation(
        data_root=args.data_root,
        output_root=args.output_root,
        run_name=args.run_name,
        years=list(args.years or []),
        workers=args.workers,
        limit=args.limit,
        input_manifest=args.input_manifest,
        include_npm1=args.include_npm1,
        dit_only=args.dit_only,
        timeout_seconds=args.timeout_seconds,
        checkpoint_every=args.checkpoint_every,
        required_run_name_contains=args.require_run_name_contains,
        excluded_basenames=list(args.excluded_basenames or DEFAULT_EXCLUDED_BASENAMES),
    )
    print(json.dumps(payload, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
