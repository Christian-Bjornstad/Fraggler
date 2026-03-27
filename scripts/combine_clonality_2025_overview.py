from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


TARGET_SHEETS = ("Patient_Runs", "Control_Runs", "PK_Peaks")


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Combine monthly 2025 clonality tracking workbooks into one overview workbook."
    )
    parser.add_argument(
        "--run-root",
        type=Path,
        required=True,
        help="Full 2025 run root containing month_runs/2025_*/track-clonality.xlsx",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Output workbook path. Defaults to <run-root>/track-clonality-2025-overview.xlsx",
    )
    return parser


def _read_month_sheet(workbook_path: Path, month: str, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(workbook_path, sheet_name=sheet_name)
    df.insert(0, "Month", month)
    return df


def _auto_fit_columns(ws) -> None:
    for column_cells in ws.columns:
        lengths = []
        for cell in column_cells:
            if cell.value is None:
                continue
            lengths.append(len(str(cell.value)))
        if not lengths:
            continue
        width = min(max(lengths) + 2, 40)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def _style_workbook(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
        ws.freeze_panes = "A2"
        _auto_fit_columns(ws)

    if "Overview" in wb.sheetnames:
        ws = wb["Overview"]
        ws.freeze_panes = "A1"
        if ws["A1"].value:
            ws["A1"].font = Font(bold=True, size=14)

    wb.save(workbook_path)


def combine_run_root(run_root: Path, output_path: Path) -> Path:
    month_runs_root = run_root / "month_runs"
    if not month_runs_root.is_dir():
        raise FileNotFoundError(f"Missing month_runs directory: {month_runs_root}")

    month_names = sorted(p.name for p in month_runs_root.iterdir() if p.is_dir())
    combined: dict[str, list[pd.DataFrame]] = {sheet: [] for sheet in TARGET_SHEETS}
    overview_rows: list[dict[str, object]] = []

    for month in month_names:
        workbook_path = month_runs_root / month / "track-clonality.xlsx"
        if not workbook_path.exists():
            continue
        xl = pd.ExcelFile(workbook_path)
        month_summary: dict[str, object] = {"Month": month}
        for sheet_name in TARGET_SHEETS:
            if sheet_name not in xl.sheet_names:
                month_summary[f"{sheet_name}_Rows"] = 0
                continue
            df = _read_month_sheet(workbook_path, month, sheet_name)
            combined[sheet_name].append(df)
            month_summary[f"{sheet_name}_Rows"] = int(len(df))
            if sheet_name == "Patient_Runs" and not df.empty:
                month_summary["Patient_LadderReviewRequired"] = int((df["LadderQC"] == "review_required").sum())
                month_summary["Patient_AutoPartial"] = int((df["LadderFitStrategy"] == "auto_partial").sum())
            if sheet_name == "PK_Peaks" and not df.empty:
                month_summary["PK_Outliers_AbsDeltaGT2"] = int((pd.to_numeric(df["AbsDeltaBP"], errors="coerce") > 2.0).sum())
        overview_rows.append(month_summary)

    overview_df = pd.DataFrame(overview_rows).fillna(0)
    patient_df = pd.concat(combined["Patient_Runs"], ignore_index=True) if combined["Patient_Runs"] else pd.DataFrame()
    control_df = pd.concat(combined["Control_Runs"], ignore_index=True) if combined["Control_Runs"] else pd.DataFrame()
    pk_df = pd.concat(combined["PK_Peaks"], ignore_index=True) if combined["PK_Peaks"] else pd.DataFrame()

    total_row = {
        "Month": "TOTAL",
        "Patient_Runs_Rows": int(len(patient_df)),
        "Control_Runs_Rows": int(len(control_df)),
        "PK_Peaks_Rows": int(len(pk_df)),
        "Patient_LadderReviewRequired": int((patient_df.get("LadderQC", pd.Series(dtype=object)) == "review_required").sum()) if not patient_df.empty else 0,
        "Patient_AutoPartial": int((patient_df.get("LadderFitStrategy", pd.Series(dtype=object)) == "auto_partial").sum()) if not patient_df.empty else 0,
        "PK_Outliers_AbsDeltaGT2": int((pd.to_numeric(pk_df.get("AbsDeltaBP", pd.Series(dtype=float)), errors="coerce") > 2.0).sum()) if not pk_df.empty else 0,
    }
    overview_df = pd.concat([overview_df, pd.DataFrame([total_row])], ignore_index=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        overview_df.to_excel(writer, sheet_name="Overview", index=False)
        patient_df.to_excel(writer, sheet_name="Patient_Runs_2025", index=False)
        control_df.to_excel(writer, sheet_name="Control_Runs_2025", index=False)
        pk_df.to_excel(writer, sheet_name="PK_Peaks_2025", index=False)

    _style_workbook(output_path)
    return output_path


def main(argv: list[str] | None = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)
    output_path = args.output or (args.run_root / "track-clonality-2025-overview.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    result = combine_run_root(args.run_root, output_path)
    print(result)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
