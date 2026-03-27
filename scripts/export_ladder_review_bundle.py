from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))


DEFAULT_MONTH_RUNS_ROOT = Path(
    "/Users/christian/Desktop/Excel_Fraggler/full_2025_runs/full_2025_validation_20260325_recovery/month_runs"
)
DEFAULT_OUTPUT_DIR = Path("/Users/christian/Desktop/Excel_Fraggler/ladder_review_bundle")


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Export a compact ladder review bundle from finished monthly clonality runs. "
            "The bundle includes only review-required or auto_partial ladder cases and the "
            "candidate ladders associated with those cases."
        )
    )
    parser.add_argument(
        "--month-runs-root",
        type=Path,
        default=DEFAULT_MONTH_RUNS_ROOT,
        help=f"Directory containing monthly run folders. Default: {DEFAULT_MONTH_RUNS_ROOT}",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help=f"Directory where the review bundle will be written. Default: {DEFAULT_OUTPUT_DIR}",
    )
    parser.add_argument(
        "--month",
        dest="months",
        action="append",
        default=[],
        help="Optional month folder name like 2025_01. Repeat to restrict export.",
    )
    return parser


def _normalize_months(values: list[str]) -> list[str]:
    normalized = [value.strip() for value in values if value.strip()]
    return normalized


def _iter_month_dirs(month_runs_root: Path, selected_months: list[str]) -> list[Path]:
    month_dirs = [p for p in sorted(month_runs_root.iterdir()) if p.is_dir() and p.name.startswith("2025_")]
    if not selected_months:
        return month_dirs
    selected = set(selected_months)
    return [month_dir for month_dir in month_dirs if month_dir.name in selected]


def _load_patient_ladder_cases(workbook_path: Path, month: str) -> list[dict[str, object]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb["Patient_Runs"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {name: i for i, name in enumerate(header)}
    exported: list[dict[str, object]] = []
    for row in rows:
        ladder_qc = row[idx["LadderQC"]]
        ladder_fit_strategy = row[idx["LadderFitStrategy"]]
        if ladder_qc != "review_required" and ladder_fit_strategy != "auto_partial":
            continue
        exported.append(
            {
                "month": month,
                "scope": "Patient",
                "identity_key": row[idx["IdentityKey"]],
                "source_run_dir": row[idx["SourceRunDir"]],
                "assay": row[idx["Assay"]],
                "run_date": row[idx["RunDate"]],
                "run_code": row[idx["RunCode"]],
                "well": row[idx["Well"]],
                "ladder": row[idx["Ladder"]],
                "ladder_qc": ladder_qc,
                "ladder_fit_strategy": ladder_fit_strategy,
                "ladder_expected_step_count": row[idx["LadderExpectedStepCount"]],
                "ladder_fitted_step_count": row[idx["LadderFittedStepCount"]],
                "ladder_r2": row[idx["LadderR2"]],
            }
        )
    return exported


def _index_ladder_features(feature_csv: Path) -> dict[tuple[str, str, str], dict[str, str]]:
    indexed: dict[tuple[str, str, str], dict[str, str]] = {}
    with feature_csv.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            if row.get("scope") != "Patient":
                continue
            key = (
                row.get("source_run_dir", ""),
                row.get("assay", ""),
                row.get("well", ""),
            )
            indexed[key] = row
    return indexed


def _load_candidate_rows(candidate_csv: Path) -> dict[tuple[str, str, str], list[dict[str, str]]]:
    grouped: dict[tuple[str, str, str], list[dict[str, str]]] = {}
    with candidate_csv.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            key = (
                row.get("source_run_dir", ""),
                row.get("assay", ""),
                row.get("well", ""),
            )
            grouped.setdefault(key, []).append(row)
    for rows in grouped.values():
        rows.sort(key=lambda row: int(row.get("candidate_index") or 0))
    return grouped


def export_ladder_review_bundle(
    month_runs_root: Path,
    output_dir: Path,
    selected_months: list[str] | None = None,
) -> dict[str, object]:
    month_runs_root = Path(month_runs_root).expanduser().resolve()
    output_dir = Path(output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    months = _iter_month_dirs(month_runs_root, selected_months or [])
    case_rows: list[dict[str, object]] = []
    candidate_rows: list[dict[str, object]] = []

    for month_dir in months:
        workbook_path = month_dir / "track-clonality.xlsx"
        feature_csv = month_dir / "feature_artifacts" / "clonality_ladder_features.csv"
        candidate_csv = month_dir / "candidate_artifacts" / "clonality_ladder_candidates.csv"
        if not (workbook_path.exists() and feature_csv.exists() and candidate_csv.exists()):
            continue

        feature_index = _index_ladder_features(feature_csv)
        candidate_index = _load_candidate_rows(candidate_csv)

        for case in _load_patient_ladder_cases(workbook_path, month_dir.name):
            key = (str(case["source_run_dir"]), str(case["assay"]), str(case["well"]))
            feature_row = feature_index.get(key, {})
            case_row = dict(case)
            case_row["artifact_row_key"] = feature_row.get("artifact_row_key", "")
            case_row["join_key"] = feature_row.get("join_key", "")
            case_row["ladder_join_key"] = feature_row.get("ladder_join_key", "")
            case_row["label"] = ""
            case_row["label_note"] = ""
            case_rows.append(case_row)

            for candidate in candidate_index.get(key, []):
                candidate_rows.append(
                    {
                        "month": month_dir.name,
                        "source_run_dir": candidate.get("source_run_dir", ""),
                        "assay": candidate.get("assay", ""),
                        "identity_key": candidate.get("identity_key", ""),
                        "run_date": candidate.get("run_date", ""),
                        "run_code": candidate.get("run_code", ""),
                        "well": candidate.get("well", ""),
                        "ladder": candidate.get("ladder", ""),
                        "artifact_row_key": candidate.get("artifact_row_key", ""),
                        "join_key": candidate.get("join_key", ""),
                        "ladder_join_key": candidate.get("ladder_join_key", ""),
                        "candidate_index": candidate.get("candidate_index", ""),
                        "candidate_time": candidate.get("candidate_time", ""),
                        "candidate_intensity": candidate.get("candidate_intensity", ""),
                        "candidate_source": candidate.get("candidate_source", ""),
                        "selected_for_fit": candidate.get("selected_for_fit", ""),
                        "selected_step_bp": candidate.get("selected_step_bp", ""),
                        "ladder_fit_strategy": candidate.get("ladder_fit_strategy", ""),
                        "ladder_r2": candidate.get("ladder_r2", ""),
                        "ladder_review_required": candidate.get("ladder_review_required", ""),
                        "human_label": "",
                        "human_note": "",
                    }
                )

    case_csv = output_dir / "ladder_review_cases.csv"
    candidate_csv = output_dir / "ladder_review_candidates.csv"
    summary_json = output_dir / "ladder_review_summary.json"

    case_fields = [
        "month",
        "scope",
        "identity_key",
        "source_run_dir",
        "assay",
        "run_date",
        "run_code",
        "well",
        "ladder",
        "ladder_qc",
        "ladder_fit_strategy",
        "ladder_expected_step_count",
        "ladder_fitted_step_count",
        "ladder_r2",
        "artifact_row_key",
        "join_key",
        "ladder_join_key",
        "label",
        "label_note",
    ]
    with case_csv.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=case_fields)
        writer.writeheader()
        writer.writerows(case_rows)

    candidate_fields = [
        "month",
        "source_run_dir",
        "assay",
        "identity_key",
        "run_date",
        "run_code",
        "well",
        "ladder",
        "artifact_row_key",
        "join_key",
        "ladder_join_key",
        "candidate_index",
        "candidate_time",
        "candidate_intensity",
        "candidate_source",
        "selected_for_fit",
        "selected_step_bp",
        "ladder_fit_strategy",
        "ladder_r2",
        "ladder_review_required",
        "human_label",
        "human_note",
    ]
    with candidate_csv.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=candidate_fields)
        writer.writeheader()
        writer.writerows(candidate_rows)

    run_counts = Counter(str(row["source_run_dir"]) for row in case_rows)
    assay_counts = Counter(str(row["assay"]) for row in case_rows)
    summary = {
        "generated_at_utc": datetime_now_utc(),
        "month_runs_root": str(month_runs_root),
        "output_dir": str(output_dir),
        "selected_months": selected_months or [],
        "case_count": len(case_rows),
        "candidate_count": len(candidate_rows),
        "run_counts": dict(run_counts),
        "assay_counts": dict(assay_counts),
        "outputs": {
            "cases_csv": str(case_csv),
            "candidates_csv": str(candidate_csv),
            "summary_json": str(summary_json),
        },
    }
    summary_json.write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return summary


def datetime_now_utc() -> str:
    from datetime import datetime, timezone

    return datetime.now(timezone.utc).isoformat()


def main(argv: list[str] | None = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)
    summary = export_ladder_review_bundle(
        args.month_runs_root,
        args.output_dir,
        _normalize_months(args.months),
    )
    print(json.dumps(summary, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
