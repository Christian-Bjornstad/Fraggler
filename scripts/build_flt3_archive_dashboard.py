from __future__ import annotations

import argparse
import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


KNOWN_ISSUE_FILES: dict[str, str] = {
    "25OUM13468_D835__050925_F04_C990WOJA.fsa": "Known missing ladder / machine issue",
    "25OUM00872_ITD__220125_B01_C9U07BS8.fsa": "Known human or run-quality issue",
    "25OUM00872_ITD-Ratio__220125_B03_C9U07BS8.fsa": "Known human or run-quality issue",
    "25OUM00872_ITD-Ratio__220125_A03_C9U07BS8.fsa": "Known human or run-quality issue",
    "25OUM01097_ITD-Ratio__220125_C03_C9U07BS8.fsa": "Known human or run-quality issue",
    "25OUM00872_ITD_forty__220125_A02_C9U07BS8.fsa": "Known human or run-quality issue",
    "25OUM04778_p2_RATIO__250324_F04_H9C0VADZ.fsa": "Known missing ladder / machine issue",
    "25OUM02961_ITD__200225_A04_C9U07BS8.fsa": "Known human or run-quality issue",
    "25OUM02961_ITD_forty__200225_D04_C9U07BS8.fsa": "Known human or run-quality issue",
    "25OUM02961_ITD-Ratio__200225_F04_C9U07BS8.fsa": "Known human or run-quality issue",
    "25OUM00872_ITD__220125_A01_C9U07BS8.fsa": "Known human or run-quality issue",
    "25OUM01097_ITD-Ratio__220125_D03_C9U07BS8.fsa": "Known human or run-quality issue",
    "25OUM11534_p2_TKD-kutting__240725_B05_H9C0VC6E.fsa": "Known human or run-quality issue",
    "25OUM04888_p2_RATIO__250324_H04_H9C0VADZ.fsa": "Known missing ladder / machine issue",
    "25OUM02961_ITD__200225_B04_C9U07BS8.fsa": "Known human or run-quality issue",
    "25OUM12881_itd-Ratio__250825_F04_C990WO66.fsa": "Known human or run-quality issue",
    "25OUM12253_RATIO__130825_A03_C990WO65.fsa": "Known missing ladder / machine issue",
    "25OUM01097_ITD_forty__220125_C02_C9U07BS8.fsa": "Known human or run-quality issue",
    "25OUM01097_ITD__220125_C01_C9U07BS8.fsa": "Known human or run-quality issue",
    "25OUM02961_ITD-Ratio__200225_E04_C9U07BS8.fsa": "Known human or run-quality issue",
    "25OUM13823_Itd-Ratio__160925_F04_C990WOCK.fsa": "Known human or run-quality issue",
    "25OUM11314_p2_RATIO__310725_F04_H9C0ZIZJ.fsa": "Known missing ladder / machine issue",
}


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build a leader-friendly FLT3 archive dashboard workbook.")
    parser.add_argument("--run-dir", type=Path, required=True, help="Completed FLT3 archive validation directory.")
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Optional output workbook path. Defaults to <run-dir>/track-flt3-archive-dashboard.xlsx.",
    )
    return parser


def _read_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    return pd.read_csv(path).fillna("")


def _safe_num(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def _dedupe_metrics(metrics: pd.DataFrame) -> pd.DataFrame:
    if metrics.empty or "path" not in metrics.columns:
        return metrics

    df = metrics.copy()
    df["status"] = df.get("status", "").astype(str)
    df["ladder_fit_strategy"] = df.get("ladder_fit_strategy", "").astype(str)
    df["ladder_r2_num"] = _safe_num(df.get("ladder_r2", pd.Series(dtype=float)))
    df["max_abs_error_bp_num"] = _safe_num(df.get("max_abs_error_bp", pd.Series(dtype=float)))

    status_rank = {
        "ok": 0,
        "manual_adjustment": 1,
        "review_required": 2,
        "ladder_fit_failed": 3,
    }
    df["status_rank"] = df["status"].map(lambda value: status_rank.get(str(value), 9))
    df["strategy_rank"] = df["ladder_fit_strategy"].map(
        lambda value: 0 if str(value) == "manual_adjustment" else 1
    )
    dedupe_key = "file" if "file" in df.columns else "path"
    df = df.sort_values(
        by=[dedupe_key, "status_rank", "strategy_rank", "max_abs_error_bp_num", "ladder_r2_num", "path"],
        ascending=[True, True, True, True, False, True],
    )
    deduped = df.drop_duplicates(subset=[dedupe_key], keep="first").copy()
    return deduped.drop(columns=["status_rank", "strategy_rank"], errors="ignore")


def _style_sheet(ws) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, column_cells in enumerate(ws.columns, start=1):
        sample = [str(cell.value) if cell.value is not None else "" for cell in column_cells[:300]]
        width = min(max(max((len(v) for v in sample), default=0) + 2, 12), 54)
        ws.column_dimensions[get_column_letter(idx)].width = width


def _build_dashboard_sheet(wb, overview_rows: list[tuple[str, object]], daily_df: pd.DataFrame, run_df: pd.DataFrame) -> None:
    ws = wb["Dashboard"]
    if ws.max_row > 1 or ws.max_column > 1 or ws["A1"].value is not None:
        ws.delete_rows(1, ws.max_row)
    ws["A1"] = "FLT3 Archive Validation Dashboard"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A3"] = "Key Metrics"
    ws["A3"].font = Font(bold=True)

    row = 4
    for metric, value in overview_rows:
        ws[f"A{row}"] = metric
        ws[f"B{row}"] = value
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

    if not daily_df.empty:
        start = row + 2
        ws[f"A{start}"] = "Runs Per Date"
        ws[f"A{start}"].font = Font(bold=True)
        table_start = start + 1
        for c_idx, col in enumerate(daily_df.columns, start=1):
            ws.cell(row=table_start, column=c_idx, value=col)
        for r_idx, (_, r) in enumerate(daily_df.iterrows(), start=table_start + 1):
            for c_idx, col in enumerate(daily_df.columns, start=1):
                ws.cell(row=r_idx, column=c_idx, value=r[col])
        chart = LineChart()
        chart.title = "Validated Files Per Run Date"
        chart.y_axis.title = "Validated"
        chart.x_axis.title = "Run Date"
        data = Reference(ws, min_col=2, min_row=table_start, max_row=table_start + len(daily_df))
        cats = Reference(ws, min_col=1, min_row=table_start + 1, max_row=table_start + len(daily_df))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 7
        chart.width = 14
        ws.add_chart(chart, f"E{start}")

    if not run_df.empty:
        start = max(ws.max_row + 3, 24)
        ws[f"A{start}"] = "Top Review Runs"
        ws[f"A{start}"].font = Font(bold=True)
        top = run_df.head(12)
        table_start = start + 1
        for c_idx, col in enumerate(top.columns, start=1):
            ws.cell(row=table_start, column=c_idx, value=col)
        for r_idx, (_, r) in enumerate(top.iterrows(), start=table_start + 1):
            for c_idx, col in enumerate(top.columns, start=1):
                ws.cell(row=r_idx, column=c_idx, value=r[col])
        chart = BarChart()
        chart.title = "Review Cases Per Run"
        chart.y_axis.title = "Review Count"
        chart.x_axis.title = "Run Dir"
        data = Reference(ws, min_col=3, min_row=table_start, max_row=table_start + len(top))
        cats = Reference(ws, min_col=1, min_row=table_start + 1, max_row=table_start + len(top))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 16
        ws.add_chart(chart, f"G{start}")


def build_dashboard_workbook(run_dir: Path, output_path: Path) -> Path:
    summary = json.loads((run_dir / "summary.json").read_text())
    metrics_raw = _read_csv(run_dir / "flt3_ladder_metrics.csv")
    metrics = _dedupe_metrics(metrics_raw)
    manual = _read_csv(run_dir / "flt3_ladder_manual_adjustments_applied.csv")
    residual_by_year = _read_csv(run_dir / "residual_by_year.csv")
    residual_by_assay = _read_csv(run_dir / "residual_by_assay.csv")
    residual_by_analysis = _read_csv(run_dir / "residual_by_analysis_type.csv")
    residual_by_strategy = _read_csv(run_dir / "residual_by_strategy.csv")
    residual_worst = _read_csv(run_dir / "residual_worst_cases.csv")

    metrics["known_issue_reason"] = metrics["file"].map(KNOWN_ISSUE_FILES).fillna("")
    metrics["known_issue"] = metrics["known_issue_reason"].ne("")
    review = metrics.loc[metrics["status"].isin(["review_required", "ladder_fit_failed"])].copy()
    review["queue_group"] = ""
    review.loc[review["status"] == "ladder_fit_failed", "queue_group"] = "fit_failed"
    review.loc[
        (review["status"] == "review_required")
        & review["review_reason"].astype(str).str.contains("Missing expected ladder steps", na=False),
        "queue_group",
    ] = "missing_steps"
    review.loc[
        (review["status"] == "review_required")
        & review["review_reason"].astype(str).str.contains("Max residual", na=False),
        "queue_group",
    ] = "high_residual"
    review["r2"] = review.get("ladder_r2", "")
    review["max_bp_err"] = review.get("max_abs_error_bp", "")
    review["mean_bp_err"] = review.get("mean_abs_error_bp", "")
    review["expected_step_count"] = review.get("ladder_expected_step_count", "")
    review["fitted_step_count"] = review.get("ladder_fitted_step_count", "")
    review["missing_expected_steps"] = review.get("ladder_missing_expected_steps", "")
    review["known_issue_reason"] = review["file"].map(KNOWN_ISSUE_FILES).fillna("")
    review["known_issue"] = review["known_issue_reason"].ne("")

    filtered_metrics = metrics.loc[~metrics["known_issue"]].copy()
    filtered_review = review.loc[~review["known_issue"]].copy()
    excluded_review = review.loc[review["known_issue"]].copy()
    filtered_fit_failed = filtered_review.loc[filtered_review["status"] == "ladder_fit_failed"].copy()
    filtered_residual_worst = residual_worst.loc[~residual_worst["file"].isin(KNOWN_ISSUE_FILES)].copy()

    validated = filtered_metrics.loc[~filtered_metrics["status"].astype(str).str.startswith("skipped")].copy()
    validated["max_abs_error_bp_num"] = _safe_num(validated.get("max_abs_error_bp", pd.Series(dtype=float)))
    validated["ladder_r2_num"] = _safe_num(validated.get("ladder_r2", pd.Series(dtype=float)))
    validated["run_date_fmt"] = validated["run_date"].astype(str).replace({"": "Unknown"})

    daily = (
        validated.groupby("run_date_fmt", dropna=False)
        .agg(
            validated_files=("file", "count"),
            review_required=("status", lambda s: int((s == "review_required").sum())),
            ladder_fit_failed=("status", lambda s: int((s == "ladder_fit_failed").sum())),
            median_max_residual_bp=("max_abs_error_bp_num", "median"),
        )
        .reset_index()
        .sort_values("run_date_fmt")
    )

    run_summary = (
        filtered_review.groupby(["run_dir", "run_date"], dropna=False)
        .agg(
            review_count=("file", "count"),
            fit_failed=("queue_group", lambda s: int((s == "fit_failed").sum())),
            high_residual=("queue_group", lambda s: int((s == "high_residual").sum())),
            missing_steps=("queue_group", lambda s: int((s == "missing_steps").sum())),
        )
        .reset_index()
        .sort_values(["review_count", "fit_failed", "high_residual"], ascending=[False, False, False])
    )

    overview_rows = [
        ("Validated rows (filtered)", int(len(validated))),
        ("OK (filtered)", int((validated["status"] == "ok").sum())),
        ("Manual adjustment (filtered)", int((validated["status"] == "manual_adjustment").sum())),
        ("Review required (filtered)", int((validated["status"] == "review_required").sum())),
        ("Ladder fit failed (filtered)", int((validated["status"] == "ladder_fit_failed").sum())),
        ("Known issues excluded", int(metrics["known_issue"].sum())),
        ("Median max residual bp (filtered)", float(validated["max_abs_error_bp_num"].median())),
        ("P95 max residual bp (filtered)", float(validated["max_abs_error_bp_num"].quantile(0.95))),
        ("Worst max residual bp (filtered)", float(validated["max_abs_error_bp_num"].max())),
        ("Median ladder R2 (filtered)", float(validated["ladder_r2_num"].median())),
        ("Review rows excluded as known issues", int(len(excluded_review))),
        ("Original manual review count", int(summary.get("manual_review_count", 0) or 0)),
        ("Raw rows before dedupe", int(len(metrics_raw))),
    ]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame(overview_rows, columns=["Metric", "Value"]).to_excel(writer, sheet_name="Dashboard", index=False)
        run_summary.to_excel(writer, sheet_name="Run_Summary", index=False)
        daily.to_excel(writer, sheet_name="Daily_Trend", index=False)
        residual_by_year.to_excel(writer, sheet_name="Residual_By_Year", index=False)
        residual_by_assay.to_excel(writer, sheet_name="Residual_By_Assay", index=False)
        residual_by_analysis.to_excel(writer, sheet_name="Residual_By_Analysis", index=False)
        residual_by_strategy.to_excel(writer, sheet_name="Residual_By_Strategy", index=False)
        filtered_review.to_excel(writer, sheet_name="Review_Actionable", index=False)
        excluded_review.to_excel(writer, sheet_name="Known_Issues_Excluded", index=False)
        filtered_fit_failed.to_excel(writer, sheet_name="Fit_Failed_Actionable", index=False)
        filtered_residual_worst.to_excel(writer, sheet_name="Worst_Residuals", index=False)
        manual.to_excel(writer, sheet_name="Manual_Adjustments", index=False)
        filtered_metrics.to_excel(writer, sheet_name="All_Metrics_Filtered", index=False)
        metrics_raw.to_excel(writer, sheet_name="All_Metrics_Raw", index=False)

    wb = load_workbook(output_path)
    for ws in wb.worksheets:
        _style_sheet(ws)
    _build_dashboard_sheet(wb, overview_rows, daily, run_summary)
    wb.save(output_path)
    return output_path


def main(argv: list[str] | None = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)
    run_dir = args.run_dir.expanduser().resolve()
    output = args.output.expanduser().resolve() if args.output else (run_dir / "track-flt3-archive-dashboard.xlsx")
    build_dashboard_workbook(run_dir, output)
    print(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
